#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
战新月报流水线编排器

提供完整的月报生产流水线支持，支持单步执行、断点续传、状态管理。
"""
import argparse
import asyncio
import json
import os
import shutil
import sys
import time
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any

import pandas as pd


@dataclass
class PipelineStepResult:
    step: str
    success: bool
    input_path: Optional[str] = None
    output_path: Optional[str] = None
    duration: float = 0.0
    error: Optional[str] = None
    records_count: int = 0
    metadata: Optional[Dict[str, Any]] = None


@dataclass
class PipelineState:
    batch_id: str
    created_at: str
    last_updated: str
    output_dir: str
    current_step: str
    steps: Dict[str, PipelineStepResult]
    config: Dict[str, Any]

    def save(self, path: Path):
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(asdict(self), f, ensure_ascii=False, indent=2)

    @staticmethod
    def load(path: Path) -> "PipelineState":
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return PipelineState(**data)


class ZhanxingPipeline:
    """战新月报流水线编排器"""

    STEPS = ["create_template", "crawl", "clean", "validate", "generate"]

    def __init__(self, output_dir: str, config: Optional[Dict[str, Any]] = None):
        self.output_dir = Path(output_dir)
        self.state_dir = self.output_dir / ".pipeline"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.state_dir.mkdir(parents=True, exist_ok=True)

        self.batch_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.config = config or {}
        self._state: Optional[PipelineState] = None

    @property
    def state_path(self) -> Path:
        return self.state_dir / f"{self.batch_id}.state.json"

    def _get_state(self) -> PipelineState:
        if self._state is None:
            if self.state_path.exists():
                self._state = PipelineState.load(self.state_path)
            else:
                self._state = PipelineState(
                    batch_id=self.batch_id,
                    created_at=datetime.now().isoformat(),
                    last_updated=datetime.now().isoformat(),
                    output_dir=str(self.output_dir),
                    current_step="",
                    steps={},
                    config=self.config,
                )
        return self._state

    def _save_state(self):
        state = self._get_state()
        state.last_updated = datetime.now().isoformat()
        state.save(self.state_path)

    def _get_step_output(self, step: str) -> Optional[Path]:
        """获取指定步骤的输出文件路径"""
        outputs = {
            "create_template": self.output_dir / f"候选池_{self.batch_id}.xlsx",
            "crawl": self.output_dir / f"候选池_全文_{self.batch_id}.xlsx",
            "clean": self.output_dir / f"候选池_摘要_{self.batch_id}.xlsx",
            "validate": self.output_dir / f"校验报告_{self.batch_id}.txt",
            "generate": self.output_dir / f"战新与未来产业月报_{self.batch_id}.docx",
        }
        return outputs.get(step)

    def run_step(self, step: str, **kwargs) -> PipelineStepResult:
        """执行单个步骤"""
        t0 = time.time()
        state = self._get_state()

        if step not in self.STEPS:
            return PipelineStepResult(
                step=step, success=False,
                error=f"未知步骤: {step}"
            )

        try:
            if step == "create_template":
                result = self._step_create_template(**kwargs)
            elif step == "crawl":
                result = self._step_crawl(**kwargs)
            elif step == "clean":
                result = self._step_clean(**kwargs)
            elif step == "validate":
                result = self._step_validate(**kwargs)
            elif step == "generate":
                result = self._step_generate(**kwargs)
            else:
                result = PipelineStepResult(step=step, success=False, error="未实现")

            result.duration = time.time() - t0
            state.steps[step] = result
            state.current_step = step if result.success else state.current_step
            self._save_state()
            return result

        except Exception as e:
            result = PipelineStepResult(
                step=step, success=False,
                error=str(e), duration=time.time() - t0
            )
            state.steps[step] = result
            self._save_state()
            raise

    # ---- 步骤实现 ----

    def _step_create_template(self, title: str = "战新与未来产业月报",
                              period: str = "", **kwargs) -> PipelineStepResult:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        output = self._get_step_output("create_template")
        wb = Workbook()
        ws = wb.active
        ws.title = "final_news"

        HEADERS = ["heading_1", "heading_2", "title", "content", "source", "date", "url"]
        ws.append(HEADERS)
        fill = PatternFill("solid", fgColor="1F4E78")
        font = Font(color="FFFFFF", bold=True)
        border = Border(bottom=Side(style="thin", color="808080"))
        for idx, name in enumerate(HEADERS, start=1):
            cell = ws.cell(1, idx)
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        widths = [24, 20, 38, 70, 18, 12, 50]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"

        for sheet_name in ["raw_fetch", "audit_log", "removed_items", "summary"]:
            wb.create_sheet(sheet_name)

        wb.save(output)
        print(f"[pipeline] create_template -> {output}")

        return PipelineStepResult(
            step="create_template", success=True,
            output_path=str(output),
            records_count=0,
            metadata={"title": title, "period": period}
        )

    def _step_crawl(self, input_excel: Optional[str] = None,
                     url_list: Optional[str] = None,
                     concurrency: int = 32, timeout: int = 20,
                     min_chars: int = 100,
                     intel: bool = False,
                     intel_db: str = "temp-data/crawler_intel.db",
                     intel_threshold: float = 0.65,
                     **kwargs) -> PipelineStepResult:
        from scripts.crawl_fulltext_to_excel import main as crawl_main

        # 确定输入：优先使用预提供的 Excel，其次使用上一步的输出
        if input_excel:
            input_path = Path(input_excel)
        else:
            prev = self._get_step_output("create_template")
            if prev and prev.exists():
                input_path = prev
            else:
                return PipelineStepResult(
                    step="crawl", success=False,
                    error="未找到输入 Excel 文件，请使用 --input-excel 参数指定"
                )

        output = self._get_step_output("crawl")

        # 构造 argv 并调用 crawl 脚本
        sys.argv = [
            "crawl", "--input", str(input_path),
            "--output", str(output),
            "--concurrency", str(concurrency),
            "--timeout", str(timeout),
            "--min-chars", str(min_chars),
        ]
        if intel:
            sys.argv.extend([
                "--intel",
                "--intel-db", str(intel_db),
                "--intel-threshold", str(intel_threshold),
            ])
        crawl_main()

        # 读取记录数
        count = 0
        try:
            df = pd.read_excel(output, sheet_name="final_news")
            count = len(df.dropna(how="all"))
        except Exception:
            pass

        return PipelineStepResult(
            step="crawl", success=True,
            input_path=str(input_path),
            output_path=str(output),
            records_count=count,
        )

    def _step_clean(self, input_excel: Optional[str] = None,
                    api_key: Optional[str] = None,
                    flash_model: str = "deepseek-v4-flash",
                    pro_model: str = "deepseek-v4-pro",
                    concurrency_flash: int = 20,
                    concurrency_pro: int = 5,
                    max_chars: int = 16000,
                    skip_flash: bool = False,
                    rule_only: bool = False,
                    **kwargs) -> PipelineStepResult:
        # 确定输入
        if input_excel:
            input_path = Path(input_excel)
        else:
            prev = self._get_step_output("crawl")
            if prev and prev.exists():
                input_path = prev
            else:
                return PipelineStepResult(
                    step="clean", success=False,
                    error="未找到输入 Excel 文件"
                )

        output = self._get_step_output("clean")
        api_key = api_key or os.getenv("DEEPSEEK_API_KEY", "")

        # 构造 argv 并调用 deepseek 脚本
        sys.argv = [
            "clean", "--input", str(input_path),
            "--output", str(output),
            "--api-key", api_key,
            "--flash-model", flash_model,
            "--pro-model", pro_model,
            "--concurrency-flash", str(concurrency_flash),
            "--concurrency-pro", str(concurrency_pro),
            "--max-input-chars", str(max_chars),
        ]
        if skip_flash:
            sys.argv.append("--skip-flash")
        if rule_only:
            sys.argv.append("--rule-only")

        from scripts.deepseek_clean_summarize_excel import main as clean_main
        clean_main()

        count = 0
        try:
            df = pd.read_excel(output, sheet_name="final_news")
            count = len(df[df["summary"].notna() & (df["summary"] != "")])
        except Exception:
            pass

        return PipelineStepResult(
            step="clean", success=True,
            input_path=str(input_path),
            output_path=str(output),
            records_count=count,
        )

    def _step_validate(self, input_excel: Optional[str] = None,
                       min_content_chars: int = 80,
                       sheet: str = "final_news", **kwargs) -> PipelineStepResult:
        if input_excel:
            input_path = Path(input_excel)
        else:
            prev = self._get_step_output("clean")
            if prev and prev.exists():
                input_path = prev
            else:
                prev = self._get_step_output("crawl")
                if prev and prev.exists():
                    input_path = prev
                else:
                    return PipelineStepResult(
                        step="validate", success=False,
                        error="未找到输入 Excel 文件"
                    )

        from scripts.validate_report_excel import main as validate_main

        output = self._get_step_output("validate")
        output.parent.mkdir(parents=True, exist_ok=True)

        from io import StringIO
        captured_output = StringIO()
        old_stdout = sys.stdout
        old_stderr = sys.stderr
        sys.stdout = captured_output
        sys.stderr = captured_output
        try:
            sys.argv = [
                "validate", "--input", str(input_path),
                "--sheet", sheet,
                "--min-content-chars", str(min_content_chars),
            ]
            exit_code = validate_main()
        finally:
            sys.stdout = old_stdout
            sys.stderr = old_stderr

        return PipelineStepResult(
            step="validate", success=(exit_code == 0),
            input_path=str(input_path),
            output_path=str(output),
        )

    def _step_generate(self, input_excel: Optional[str] = None,
                       template: str = "zhan_xing",
                       validate: bool = False,
                       config_path: Optional[str] = None, **kwargs) -> PipelineStepResult:
        if input_excel:
            input_path = Path(input_excel)
        else:
            prev = self._get_step_output("clean")
            if prev and prev.exists():
                input_path = prev
            else:
                prev = self._get_step_output("crawl")
                if prev and prev.exists():
                    input_path = prev
                else:
                    return PipelineStepResult(
                        step="generate", success=False,
                        error="未找到输入 Excel 文件"
                    )

        output = self._get_step_output("generate")
        output.parent.mkdir(parents=True, exist_ok=True)

        if config_path is None:
            project_root = Path(__file__).resolve().parents[2]
            config_path = str(project_root / "templates_config.yaml")

        sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "src"))
        from csv_word_converter.core import UniversalDocumentGenerator

        # 读取 Excel
        df = pd.read_excel(input_path, sheet_name="final_news")
        df = df.dropna(how="all")
        if "heading_2" in df.columns:
            df = df.dropna(subset=["heading_2"])

        for col in ["heading_1", "heading_2", "title", "content", "source", "date", "url"]:
            if col not in df.columns:
                df[col] = ""

        data = df.to_dict("records")
        print(f"[pipeline] generate: {len(data)} records")

        gen = UniversalDocumentGenerator(template, config_path=config_path)
        temp_output = gen.generate_document(data)
        shutil.copy2(temp_output, output)

        return PipelineStepResult(
            step="generate", success=True,
            input_path=str(input_path),
            output_path=str(output),
            records_count=len(data),
        )

    def run_all(self, start_from: str = "create_template", **kwargs) -> Dict[str, PipelineStepResult]:
        """一键执行全部步骤"""
        results = {}
        start_idx = self.STEPS.index(start_from) if start_from in self.STEPS else 0

        for step in self.STEPS[start_idx:]:
            print(f"\n{'='*50}")
            print(f"[pipeline] 执行步骤: {step}")
            print(f"{'='*50}")
            result = self.run_step(step, **kwargs)
            results[step] = result

            if not result.success:
                print(f"[pipeline] 步骤 {step} 失败: {result.error}")
                print(f"[pipeline] 流水线中断。可修复后使用 --resume 继续")
                break

            print(f"[pipeline] 步骤 {step} 完成 ({result.duration:.1f}s)")

        return results

    def run(self, step: Optional[str] = None, resume: bool = False,
            start_from: Optional[str] = None, **kwargs) -> Dict[str, PipelineStepResult]:
        """运行流水线，支持单步或全量"""
        if resume and self.state_path.exists():
            state = self._get_state()
            start_from = state.current_step
            print(f"[pipeline] 从断点恢复，上次执行到: {start_from}")

        if step:
            result = self.run_step(step, **kwargs)
            return {step: result}
        elif start_from:
            return self.run_all(start_from=start_from, **kwargs)
        else:
            return self.run_all(**kwargs)


def main() -> int:
    parser = argparse.ArgumentParser(
        prog="csv2word zhanxing",
        description="战新与未来产业月报流水线",
    )
    parser.add_argument(
        "--output-dir", "-o", default="./outputs",
        help="输出目录（默认: ./outputs）"
    )
    parser.add_argument(
        "--resume", action="store_true",
        help="从上次中断的步骤继续"
    )
    subparsers = parser.add_subparsers(dest="command", help="子命令")

    # create-template
    sp = subparsers.add_parser("create-template", help="创建 Excel 候选池模板")
    sp.add_argument("--output", "-o", help="输出文件路径（默认自动生成）")
    sp.add_argument("--title", default="战新与未来产业月报")
    sp.add_argument("--period", help="期号，如 2026-04")

    # crawl
    sp = subparsers.add_parser("crawl", help="抓取 URL 全文")
    sp.add_argument("--input", "-i", help="输入 Excel 文件路径")
    sp.add_argument("--output", "-o", help="输出 Excel 文件路径（默认自动生成）")
    sp.add_argument("--concurrency", "-c", type=int, default=32)
    sp.add_argument("--timeout", "-t", type=int, default=20)
    sp.add_argument("--min-chars", type=int, default=100)
    sp.add_argument("--intel", action="store_true",
                    help="启用自学习模块（记录质量日志，阈值触发 LLM 分析）")
    sp.add_argument("--intel-db", default="temp-data/crawler_intel.db",
                    help="知识库路径 (default: temp-data/crawler_intel.db)")
    sp.add_argument("--intel-threshold", type=float, default=0.65,
                    help="LLM 分析触发阈值 (default: 0.65)")

    # clean
    sp = subparsers.add_parser("clean", help="LLM 两阶段清洗摘要")
    sp.add_argument("--input", "-i", help="输入 Excel 文件路径")
    sp.add_argument("--output", "-o", help="输出 Excel 文件路径（默认自动生成）")
    sp.add_argument("--api-key", help="DeepSeek API key（默认读取 DEEPSEEK_API_KEY）")
    sp.add_argument("--flash-model", default="deepseek-v4-flash")
    sp.add_argument("--pro-model", default="deepseek-v4-pro")
    sp.add_argument("--concurrency-flash", type=int, default=20)
    sp.add_argument("--concurrency-pro", type=int, default=5)
    sp.add_argument("--max-chars", type=int, default=16000)
    sp.add_argument("--skip-flash", action="store_true")
    sp.add_argument("--rule-only", action="store_true")

    # validate
    sp = subparsers.add_parser("validate", help="校验 Excel 完整性")
    sp.add_argument("--input", "-i", help="输入 Excel 文件路径")
    sp.add_argument("--sheet", default="final_news")
    sp.add_argument("--min-content-chars", type=int, default=80)

    # generate
    sp = subparsers.add_parser("generate", help="生成 Word 月报")
    sp.add_argument("--input", "-i", help="输入 Excel 文件路径")
    sp.add_argument("--output", "-o", help="输出 docx 文件路径（默认自动生成）")
    sp.add_argument("--template", "-t", default="zhan_xing")
    sp.add_argument("--config", help="templates_config.yaml 路径")
    sp.add_argument("--validate", action="store_true")

    # all
    sp = subparsers.add_parser("all", help="一键执行全部步骤")
    sp.add_argument("--input-excel", help="已有候选池 Excel（跳过 create-template）")
    sp.add_argument("--url-list", help="URL 列表文件（跳过 crawl）")
    sp.add_argument("--concurrency", "-c", type=int, default=32)
    sp.add_argument("--api-key", help="DeepSeek API key")
    sp.add_argument("--template", "-t", default="zhan_xing")
    sp.add_argument("--skip-clean", action="store_true", help="跳过 clean 步骤")
    sp.add_argument("--skip-validate", action="store_true", help="跳过 validate 步骤")
    sp.add_argument("--intel", action="store_true",
                    help="爬虫步骤启用自学习模块")
    sp.add_argument("--intel-db", default="temp-data/crawler_intel.db",
                    help="知识库路径")
    sp.add_argument("--intel-threshold", type=float, default=0.65,
                    help="LLM 分析触发阈值")

    args = parser.parse_args()
    output_dir = getattr(args, "output_dir", "./outputs")

    pipeline = ZhanxingPipeline(output_dir=output_dir)

    if args.command == "create-template":
        result = pipeline.run_step("create_template",
                                   title=getattr(args, "title", ""),
                                   period=getattr(args, "period", ""))
        print(f"完成: {result.output_path}")
        return 0 if result.success else 1

    elif args.command == "crawl":
        result = pipeline.run_step("crawl",
                                   input_excel=getattr(args, "input", None),
                                   concurrency=getattr(args, "concurrency", 32),
                                   timeout=getattr(args, "timeout", 20),
                                   min_chars=getattr(args, "min_chars", 100),
                                   intel=getattr(args, "intel", False),
                                   intel_db=getattr(args, "intel_db", "temp-data/crawler_intel.db"),
                                   intel_threshold=getattr(args, "intel_threshold", 0.65))
        print(f"完成: {result.output_path}")
        return 0 if result.success else 1

    elif args.command == "clean":
        result = pipeline.run_step("clean",
                                   input_excel=getattr(args, "input", None),
                                   api_key=getattr(args, "api_key", None),
                                   flash_model=getattr(args, "flash_model", "deepseek-v4-flash"),
                                   pro_model=getattr(args, "pro_model", "deepseek-v4-pro"),
                                   concurrency_flash=getattr(args, "concurrency_flash", 20),
                                   concurrency_pro=getattr(args, "concurrency_pro", 5),
                                   max_chars=getattr(args, "max_chars", 16000),
                                   skip_flash=getattr(args, "skip_flash", False),
                                   rule_only=getattr(args, "rule_only", False))
        print(f"完成: {result.output_path}")
        return 0 if result.success else 1

    elif args.command == "validate":
        result = pipeline.run_step("validate",
                                   input_excel=getattr(args, "input", None),
                                   min_content_chars=getattr(args, "min_content_chars", 80),
                                   sheet=getattr(args, "sheet", "final_news"))
        print(f"完成: {result.output_path}")
        return 0 if result.success else 1

    elif args.command == "generate":
        result = pipeline.run_step("generate",
                                   input_excel=getattr(args, "input", None),
                                   template=getattr(args, "template", "zhan_xing"),
                                   config_path=getattr(args, "config", None),
                                   validate=getattr(args, "validate", False))
        print(f"完成: {result.output_path}")
        return 0 if result.success else 1

    elif args.command == "all":
        steps_to_run = ["create_template"]
        if getattr(args, "input_excel", None):
            # 跳过 create_template，从 crawl 开始
            steps_to_run = ["crawl"]

        if "crawl" in steps_to_run or not getattr(args, "skip_clean", False):
            if "crawl" not in steps_to_run:
                steps_to_run.append("crawl")
        if not getattr(args, "skip_clean", False):
            if "clean" not in steps_to_run:
                steps_to_run.append("clean")
        if not getattr(args, "skip_validate", False):
            if "validate" not in steps_to_run:
                steps_to_run.append("validate")
        if "generate" not in steps_to_run:
            steps_to_run.append("generate")

        all_kwargs = {
            "input_excel": getattr(args, "input_excel", None),
            "concurrency": getattr(args, "concurrency", 32),
            "api_key": getattr(args, "api_key", None),
            "template": getattr(args, "template", "zhan_xing"),
            "intel": getattr(args, "intel", False),
            "intel_db": getattr(args, "intel_db", "temp-data/crawler_intel.db"),
            "intel_threshold": getattr(args, "intel_threshold", 0.65),
        }

        start_from = steps_to_run[0]
        results = pipeline.run(start_from=start_from, **all_kwargs)
        for step, result in results.items():
            status = "成功" if result.success else f"失败: {result.error}"
            print(f"  {step}: {status} ({result.duration:.1f}s)")
        return 0 if all(r.success for r in results.values()) else 1

    else:
        parser.print_help()
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
